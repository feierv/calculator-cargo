# -*- coding: utf-8 -*-
"""Generează LCL_RAIL_RATES_SMILE_COMPLETE.xlsx — rulează din rădăcina proiectului."""
import json
import os
import sys

# openpyxl din vendor local
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "tmp_openpyxl"))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

META = {
    "title": "LCL RAIL RATE FROM SMILE (complet + estimări)",
    "validity": "ETD 01-28 FEB Via CHENGDU — exemplu din grilă furnizor",
    "unit": "USD/CBM",
    "note": "Prag densitate în grilă SMILE: <300 kg/CBM vs >300 kg/CBM (+7 USD/CBM peste prag). Chengdu: același tarif ambele categorii.",
}

# (Warsaw_lo, Warsaw_hi, BHD_lo, BHD_hi, Bucharest_lo, Bucharest_hi)
TIERS = {
    "CHENGDU": (105, 105, 120, 120, 165, 165),
    "YELLOW": (125, 132, 140, 147, 185, 192),
    "ORANGE": (130, 137, 145, 152, 190, 197),
    "BLUE": (132, 139, 147, 154, 192, 199),
    "URUMQI_EST": (175, 182, 190, 197, 235, 242),
}

# Mapare oraș → tier (conform grilei + completări estimative)
CITY_TIER = {
    # Din grilă SMILE direct
    "Beijing": "BLUE",
    "Changsha": "YELLOW",
    "Changzhou": "ORANGE",
    "Chengdu": "CHENGDU",
    "Chongqing": "YELLOW",
    "Fuzhou": "ORANGE",
    "Guangzhou": "YELLOW",
    "Hangzhou": "ORANGE",
    "Nanjing": "ORANGE",
    "Nantong": "ORANGE",
    "Ningbo": "YELLOW",
    "Qingdao": "ORANGE",
    "Shanghai": "YELLOW",
    "Shenzhen": "YELLOW",
    "Tianjin": "ORANGE",
    "Wenzhou": "ORANGE",
    "Wuhan": "YELLOW",
    "Xiamen": "BLUE",
    "Xian": "YELLOW",
    "Yiwu": "YELLOW",
    "Zhengzhou": "YELLOW",
    # Completări estimative (hub logistic apropiat)
    "Baoding": "BLUE",
    "Dongguan": "YELLOW",
    "Foshan": "YELLOW",
    "Haikou": "BLUE",
    "Hong Kong": "YELLOW",
    "Huizhou": "YELLOW",
    "Jinhua": "YELLOW",
    "Jining": "ORANGE",
    "Kunming": "YELLOW",
    "Macau": "YELLOW",
    "Shijiazhuang": "BLUE",
    "Sanya": "BLUE",
    "Yongkang": "YELLOW",
    "Zhongshan": "YELLOW",
    "Dongchong": "YELLOW",
    "Qingxi": "YELLOW",
    "Xiangqiao": "YELLOW",
    "Xingfu": "YELLOW",
    "Xingyang": "YELLOW",
    "Urumqi": "URUMQI_EST",
}

TIER_LABEL = {
    "CHENGDU": "Chengdu (grilă)",
    "YELLOW": "Galben — grilă SMILE",
    "ORANGE": "Portocaliu — grilă SMILE",
    "BLUE": "Albastru — grilă SMILE",
    "URUMQI_EST": "Estimat (vest CN — nu e în grilă SMILE)",
}

REGION_NAMES = {
    "Anhui",
    "Fujian",
    "Gansu",
    "Guangdong",
    "Guangxi",
    "Guizhou",
    "Hebei",
    "Heilongjiang",
    "Henan",
    "Hubei",
    "Hunan",
    "Inner Mongolia",
    "Jiangsu",
    "Jiangxi",
    "Jilin",
    "Liaoning",
    "Ningxia",
    "Qinghai",
    "Shaanxi",
    "Shandong",
    "Sichuan",
    "Tibet",
    "Xinjiang",
    "Yunnan",
    "Zhejiang",
    "Hainan",
}

# Estimare agregată pe provincie (tier reprezentativ pentru hub-uri majore din zonă)
REGION_TIER = {
    "Anhui": "ORANGE",  # Hefei — grilă portocaliu
    "Fujian": "ORANGE",  # Fuzhou / coridor est
    "Gansu": "YELLOW",
    "Guangdong": "YELLOW",  # PRD
    "Guangxi": "YELLOW",
    "Guizhou": "YELLOW",
    "Hebei": "BLUE",  # coridor Beijing
    "Heilongjiang": "ORANGE",
    "Henan": "YELLOW",  # Zhengzhou
    "Hubei": "YELLOW",  # Wuhan
    "Hunan": "YELLOW",  # Changsha
    "Inner Mongolia": "ORANGE",
    "Jiangsu": "ORANGE",  # Nanjing / est
    "Jiangxi": "YELLOW",
    "Jilin": "ORANGE",
    "Liaoning": "ORANGE",
    "Ningxia": "YELLOW",
    "Qinghai": "YELLOW",
    "Shaanxi": "YELLOW",  # Xi'an
    "Shandong": "ORANGE",  # Qingdao / Jinan
    "Sichuan": "CHENGDU",  # hub Chengdu
    "Tibet": "YELLOW",  # estimare grosieră
    "Xinjiang": "URUMQI_EST",
    "Yunnan": "YELLOW",  # Kunming
    "Zhejiang": "ORANGE",  # Hangzhou / Ningbo — medie spre portocaliu
    "Hainan": "BLUE",  # insular, ca Haikou
}

SMILE_DIRECT = frozenset(
    {
        "Beijing",
        "Changsha",
        "Changzhou",
        "Chengdu",
        "Chongqing",
        "Fuzhou",
        "Guangzhou",
        "Hangzhou",
        "Nanjing",
        "Nantong",
        "Ningbo",
        "Qingdao",
        "Shanghai",
        "Shenzhen",
        "Tianjin",
        "Wenzhou",
        "Wuhan",
        "Xiamen",
        "Xian",
        "Yiwu",
        "Zhengzhou",
    }
)


def iter_plugin_rate_rows(cities):
    """Rânduri pentru foaia Excel / tabelul Markdown — o singură sursă de adevăr."""
    for city in sorted(cities, key=lambda x: x.lower()):
        if city in REGION_NAMES:
            tier = REGION_TIER.get(city)
            if not tier:
                yield {
                    "city": city,
                    "tip": "?",
                    "grup": "—",
                    "rates": None,
                    "obs": "Lipsește REGION_TIER.",
                }
                continue
            w_lo, w_hi, b_lo, b_hi, bu_lo, bu_hi = TIERS[tier]
            yield {
                "city": city,
                "tip": "Estimativ (provincie)",
                "grup": TIER_LABEL.get(tier, tier),
                "rates": (w_lo, w_hi, b_lo, b_hi, bu_lo, bu_hi),
                "obs": "Estimare agregată pentru provincie (fără depozit anume în grilă).",
            }
            continue
        if city == "Taiwan":
            w_lo, w_hi, b_lo, b_hi, bu_lo, bu_hi = TIERS["YELLOW"]
            yield {
                "city": city,
                "tip": "Estimativ",
                "grup": TIER_LABEL["YELLOW"],
                "rates": (w_lo, w_hi, b_lo, b_hi, bu_lo, bu_hi),
                "obs": "Estimativ; flux/documentație pot diferi de coridorul China–EU standard.",
            }
            continue

        tier = CITY_TIER.get(city)
        if not tier:
            yield {
                "city": city,
                "tip": "?",
                "grup": "—",
                "rates": None,
                "obs": "Lipsește mapare — adăugați manual.",
            }
            continue

        w_lo, w_hi, b_lo, b_hi, bu_lo, bu_hi = TIERS[tier]
        src = "Grilă SMILE" if city in SMILE_DIRECT else "Estimativ"
        obs = ""
        if src == "Estimativ":
            obs = TIER_LABEL.get(tier, "") + " — hub apropiat / coridor feroviar."
        if tier == "URUMQI_EST":
            obs = "Estimat vest CN; nu e în grila SMILE."
        yield {
            "city": city,
            "tip": src,
            "grup": TIER_LABEL.get(tier, tier),
            "rates": (w_lo, w_hi, b_lo, b_hi, bu_lo, bu_hi),
            "obs": obs,
        }


def write_markdown_md(path, cities):
    lines = [
        "# LCL Rail — tarife SMILE (referință + completare pentru lista din plugin)",
        "",
        "**Unitate:** USD/CBM  ",
        "**Exemplu valabilitate (din grilă furnizor):** ETD 01–28 feb, via Chengdu  ",
        "**Notă prag densitate (SMILE):** &lt;300 kg/CBM vs &gt;300 kg/CBM — diferență **+7 USD/CBM** între cele două coloane (unde există două valori). **Chengdu:** același tarif pentru ambele categorii.",
        "",
        "Provinciile au **estimare agregată** (tier reprezentativ); nu este nevoie de un oraș concret din foaia de tarife.",
        "",
        "Acest fișier este generat de `scripts/build_lcl_rail_rates_xlsx.py` (aceeași logică ca Excel-ul).",
        "",
        "---",
        "",
        "## Ghid — legătura cu estimatorul din plugin",
        "",
        "### Ce este acest tabel",
        "",
        "- **Secțiunea 1** reproduce grupele din documentul furnizor **SMILE** (Varșovia, BHD, București) în **USD/CBM**.",
        "- **Secțiunea 2** aliniază fiecare intrare din `assets/data/cn-cities.json` la un **grup tarifar** (Chengdu / Galben / Portocaliu / Albastru sau estimare Urumqi), ca referință pentru ofertare.",
        "",
        "### Cum calculează pluginul prețul feroviar (rezumat)",
        "",
        "Implementare: `assets/js/calculator.js` — `computeRailTransportPriceEur` (metodă „pe hârtie”). Detaliu: **`RAIL_LCL_CALCULATION.md`**.",
        "",
        "| Element | Regulă în cod (estimativ) |",
        "|--------|---------------------------|",
        "| CBM taxabil | `max(m³, kg ÷ 300)` — **fără** coeficient 333 kg/m³. |",
        "| Rail (USD) | `CBM taxabil ×` tarif București (ușor/dens după densitate vs 300 kg/m³); mapare `RAIL_BUCHAREST_USD_PER_CBM`. |",
        "| Pick-up China (USD) | `CBM taxabil ×` tarif USD/CBM; default **30**; `RAIL_PICKUP_USD_PER_CBM`. |",
        "| Local (USD) | `CBM taxabil × 10 + 50` (`RAIL_LOCAL_*`). |",
        "| Extra (USD) | `CBM taxabil × 10` (`RAIL_EXTRA_USD_PER_CBM`) — ca în PASUL 6. |",
        "| Total EUR | `(rail + pickup + local + extra) USD × RAIL_USD_TO_EUR` (0,92). |",
        "| Rutier România | **Nu** este inclus la feroviar; la **aer** da. |",
        "",
        "",
        "---",
        "",
        "## 1. Grile de referință (document SMILE)",
        "",
        "| Grup | Orașe în document | Varșovia &lt;300 | Varșovia &gt;300 | BHD &lt;300 | BHD &gt;300 | București &lt;300 | București &gt;300 |",
        "|------|-------------------|------------------|------------------|-------------|------------|-------------------|------------------|",
        "| **Chengdu** | Chengdu | 105 | 105 | 120 | 120 | 165 | 165 |",
        "| **Galben** | Shenzhen, Guangzhou, Chongqing, Ningbo, Shanghai, Changsha, Xi’an, Zhengzhou, Wuhan, Yiwu | 125 | 132 | 140 | 147 | 185 | 192 |",
        "| **Portocaliu** | Changzhou, Qingdao, Tianjin, Wenzhou, Nanjing, Yangzhou, Nantong, Suzhou, Hangzhou, Wuxi, Hefei, Fuzhou | 130 | 137 | 145 | 152 | 190 | 197 |",
        "| **Albastru** | Shantou, Xiamen, Beijing | 132 | 139 | 147 | 154 | 192 | 199 |",
        "",
        "*BHD = Budapesta / Hamburg / Duisburg.*",
        "",
        "---",
        "",
        "## 2. Toate intrările din `cn-cities.json` (plugin)",
        "",
        "| Oraș | Sursă | Grup tarifar | W &lt;300 | W &gt;300 | BHD &lt;300 | BHD &gt;300 | Buc &lt;300 | Buc &gt;300 | Observații |",
        "|------|-------|--------------|-----------|-----------|-------------|------------|-------------|------------|------------|",
    ]
    for r in iter_plugin_rate_rows(cities):
        if r["rates"] is None:
            lines.append(
                "| {city} | {tip} | {grup} | — | — | — | — | — | — | {obs} |".format(**r)
            )
        else:
            w = r["rates"]
            lines.append(
                "| {city} | {tip} | {grup} | {w0} | {w1} | {w2} | {w3} | {w4} | {w5} | {obs} |".format(
                    city=r["city"],
                    tip=r["tip"],
                    grup=r["grup"],
                    w0=w[0],
                    w1=w[1],
                    w2=w[2],
                    w3=w[3],
                    w4=w[4],
                    w5=w[5],
                    obs=r["obs"] or "—",
                )
            )
    lines.extend(
        [
            "",
            "---",
            "",
            "## Fișiere înrudite",
            "",
            "| Fișier | Rol |",
            "|--------|-----|",
            "| `assets/data/LCL_RAIL_RATES_SMILE_COMPLETE.xlsx` | Aceleași date în Excel |",
            "| `scripts/build_lcl_rail_rates_xlsx.py` | Regenerare xlsx + acest fișier `.md` |",
            "| `assets/js/calculator.js` | `RAIL_BUCHAREST_USD_PER_CBM`, `computeRailTransportPriceEur` |",
            "| `RAIL_LCL_CALCULATION.md` | Formule estimator feroviar (plugin) |",
            "",
        ]
    )
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def main():
    cities_path = os.path.join(ROOT, "assets", "data", "cn-cities.json")
    with open(cities_path, encoding="utf-8") as f:
        cities = json.load(f)

    wb = Workbook()

    # --- Sheet 1: Info ---
    ws0 = wb.active
    ws0.title = "Info"
    ws0["A1"] = META["title"]
    ws0["A1"].font = Font(bold=True, size=14)
    ws0["A3"] = "Valabilitate (exemplu grilă):"
    ws0["B3"] = META["validity"]
    ws0["A4"] = "Unitate:"
    ws0["B4"] = META["unit"]
    ws0["A5"] = "Notă:"
    ws0["B5"] = META["note"]
    ws0["A7"] = "Grupuri în documentul SMILE original:"
    rows_info = [
        ("Chengdu", "105", "120", "165", "același tarif ambele categorii greutate"),
        ("Galben: Shenzhen, Guangzhou, Chongqing, Ningbo, Shanghai, Changsha, Xi'an, Zhengzhou, Wuhan, Yiwu", "125/132", "140/147", "185/192", ""),
        ("Portocaliu: Changzhou, Qingdao, Tianjin, Wenzhou, Nanjing, Yangzhou, Nantong, Suzhou, Hangzhou, Wuxi, Hefei, Fuzhou", "130/137", "145/152", "190/197", ""),
        ("Albastru: Shantou, Xiamen, Beijing", "132/139", "147/154", "192/199", ""),
    ]
    r = 8
    for line in rows_info:
        for c, val in enumerate(line, 1):
            ws0.cell(row=r, column=c, value=val)
        r += 1

    # --- Sheet 2: Grile referință ---
    ws_ref = wb.create_sheet("Grile_SMILE_referinta")
    hdr = [
        "Grup",
        "Orașe (exemplu document)",
        "Warsaw_PL USD <300kg/CBM",
        "Warsaw_PL USD >300kg/CBM",
        "Budapest_Hamburg_Duisburg USD <300",
        "Budapest_Hamburg_Duisburg USD >300",
        "Bucharest USD <300",
        "Bucharest USD >300",
    ]
    for col, h in enumerate(hdr, 1):
        c = ws_ref.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDEBF7")

    ref_data = [
        ("CHENGDU", "Chengdu", 105, 105, 120, 120, 165, 165),
        ("YELLOW", "Shenzhen, Guangzhou, Chongqing, Ningbo, Shanghai, Changsha, Xi'an, Zhengzhou, Wuhan, Yiwu", 125, 132, 140, 147, 185, 192),
        ("ORANGE", "Changzhou, Qingdao, Tianjin, Wenzhou, Nanjing, Yangzhou, Nantong, Suzhou, Hangzhou, Wuxi, Hefei, Fuzhou", 130, 137, 145, 152, 190, 197),
        ("BLUE", "Shantou, Xiamen, Beijing", 132, 139, 147, 154, 192, 199),
    ]
    for i, row in enumerate(ref_data, 2):
        for j, val in enumerate(row, 1):
            ws_ref.cell(row=i, column=j, value=val)

    # --- Sheet 3: Toate orașele din plugin ---
    ws = wb.create_sheet("Orase_plugin_complet")
    headers = [
        "Oraș",
        "Tip",
        "Grup tarifar",
        "Warsaw <300",
        "Warsaw >300",
        "Budapest_Hamburg_Duisburg <300",
        "Budapest_Hamburg_Duisburg >300",
        "Bucharest <300",
        "Bucharest >300",
        "Observații",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="FFF2CC")

    row = 2
    for r in iter_plugin_rate_rows(cities):
        ws.cell(row=row, column=1, value=r["city"])
        ws.cell(row=row, column=2, value=r["tip"])
        ws.cell(row=row, column=3, value=r["grup"])
        if r["rates"]:
            w_lo, w_hi, b_lo, b_hi, bu_lo, bu_hi = r["rates"]
            ws.cell(row=row, column=4, value=w_lo)
            ws.cell(row=row, column=5, value=w_hi)
            ws.cell(row=row, column=6, value=b_lo)
            ws.cell(row=row, column=7, value=b_hi)
            ws.cell(row=row, column=8, value=bu_lo)
            ws.cell(row=row, column=9, value=bu_hi)
        ws.cell(row=row, column=10, value=r["obs"])
        row += 1

    # Lățimi coloane
    for sheet in (ws_ref, ws):
        for col in range(1, 12):
            sheet.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["J"].width = 45
    ws_ref.column_dimensions["B"].width = 55

    out_path = os.path.join(ROOT, "assets", "data", "LCL_RAIL_RATES_SMILE_COMPLETE.xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print("Saved:", out_path)

    md_path = os.path.join(ROOT, "assets", "data", "LCL_RAIL_RATES_SMILE_COMPLETE.md")
    write_markdown_md(md_path, cities)
    print("Saved:", md_path)


if __name__ == "__main__":
    main()
